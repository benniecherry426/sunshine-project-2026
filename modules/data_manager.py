import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
import streamlit as st
from config import SPREADSHEET_FOLDER

class DataManager:
    def __init__(self):
        self.spreadsheet_folder = SPREADSHEET_FOLDER
        os.makedirs(self.spreadsheet_folder, exist_ok=True)
    
    def create_spreadsheet(self, spreadsheet_name, categories):
        """
        Create a new spreadsheet with specified categories
        """
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Financial Data"
            
            # Create header
            headers = ["Date", "Description", "Amount"] + categories
            ws.append(headers)
            
            # Style header
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Save spreadsheet
            file_path = os.path.join(self.spreadsheet_folder, f"{spreadsheet_name}.xlsx")
            wb.save(file_path)
            st.success(f"✅ Spreadsheet '{spreadsheet_name}' created successfully!")
            return file_path
            
        except Exception as e:
            st.error(f"❌ Error creating spreadsheet: {str(e)}")
            return None
    
    def add_data_to_spreadsheet(self, file_path, data_dict):
        """
        Add data to existing spreadsheet
        data_dict: Dictionary with columns as keys
        """
        try:
            df = pd.read_excel(file_path)
            new_row = pd.DataFrame([data_dict])
            df = pd.concat([df, new_row], ignore_index=True)
            df.to_excel(file_path, index=False)
            st.success("✅ Data added to spreadsheet!")
            return True
        except Exception as e:
            st.error(f"❌ Error adding data: {str(e)}")
            return False
    
    def get_spreadsheet_data(self, file_path):
        """
        Read and return spreadsheet data
        """
        try:
            df = pd.read_excel(file_path)
            return df
        except Exception as e:
            st.error(f"❌ Error reading spreadsheet: {str(e)}")
            return None
    
    def list_spreadsheets(self):
        """
        List all available spreadsheets
        """
        try:
            files = [f for f in os.listdir(self.spreadsheet_folder) if f.endswith('.xlsx')]
            return files
        except Exception as e:
            st.error(f"❌ Error listing spreadsheets: {str(e)}")
            return []
    
    def parse_financial_data(self, text, categories):
        """
        Parse financial data from text using AI
        """
        import openai
        from config import OPENAI_API_KEY, OPENAI_MODEL
        
        openai.api_key = OPENAI_API_KEY
        
        try:
            prompt = f"""
            Extract financial data from the following text and organize it by these categories: {', '.join(categories)}
            
            Return the data in JSON format with keys: date, description, amount, and category.
            
            Text: {text}
            """
            
            response = openai.ChatCompletion.create(
                model=OPENAI_MODEL,
                messages=[{"role": "user", "content": prompt}]
            )
            
            return response['choices'][0]['message']['content']
        except Exception as e:
            st.error(f"❌ Error parsing data: {str(e)}")
            return None
